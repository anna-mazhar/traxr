"""Artifact loading for the v1 modalities (CSV/XLSX/TXT/MD/PDF).

Design notes:

* dict-style ``{'success': False, 'error': ...}`` results became the typed
  error hierarchy (:class:`~traxr.errors.InvalidArtifactError`,
  :class:`~traxr.errors.UnsupportedModalityError`,
  :class:`~traxr.errors.ModalityMismatchError`,
  :class:`~traxr.errors.OptionalDependencyError`);
* trimmed to v1 scope — TABULAR = CSV/XLSX, DOCUMENT = PDF/TXT/MD. docx/pptx
  detection remains but maps to ``UnsupportedModalityError`` with a "coming in
  a future release" message; image/audio/archive likewise.
"""

import csv
import mimetypes
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from traxr.errors import (
    InvalidArtifactError,
    ModalityMismatchError,
    OptionalDependencyError,
    UnsupportedModalityError,
)

# Optional imports (module-level flags so tests can simulate missing deps).
try:
    import fitz  # PyMuPDF

    HAS_PYMUPDF = True
except ImportError:  # pragma: no cover - environment-dependent
    HAS_PYMUPDF = False

try:
    import pdfplumber

    HAS_PDFPLUMBER = True
except ImportError:  # pragma: no cover - environment-dependent
    HAS_PDFPLUMBER = False

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - environment-dependent
    HAS_OPENPYXL = False


SUPPORTED_SUMMARY = "CSV/XLSX/TXT/MD/PDF"

# v1 extension sets (category sets inherited from FileHandler/FileInspector).
TABULAR_EXTENSIONS = {".csv", ".xlsx"}
DOCUMENT_EXTENSIONS = {".pdf", ".txt", ".md"}

# Known-but-unsupported categories, for actionable error messages.
FUTURE_DOCUMENT_EXTENSIONS = {".docx", ".doc", ".pptx", ".ppt"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff"}
AUDIO_EXTENSIONS = {".mp3", ".wav", ".ogg", ".flac", ".m4a", ".aac"}
ARCHIVE_EXTENSIONS = {".zip", ".tar", ".gz", ".7z", ".rar"}

# Content signatures used to catch wrong-extension artifacts.
_PDF_MAGIC = b"%PDF"
_ZIP_MAGIC = b"PK\x03\x04"  # xlsx/docx/pptx are zip containers


def file_type_for(path: str | Path) -> str:
    """Map a path to its v1 file type (``csv``/``xlsx``/``txt``/``md``/``pdf``).

    Raises:
        UnsupportedModalityError: for any other extension.
    """
    ext = Path(path).suffix.lower()
    if ext in TABULAR_EXTENSIONS or ext in DOCUMENT_EXTENSIONS:
        return ext.lstrip(".")
    if ext in FUTURE_DOCUMENT_EXTENSIONS:
        raise UnsupportedModalityError(
            f"'{ext}' documents are not supported yet (coming in a future "
            f"release). v1 supports {SUPPORTED_SUMMARY}."
        )
    if ext in IMAGE_EXTENSIONS or ext in AUDIO_EXTENSIONS:
        raise UnsupportedModalityError(
            f"'{ext}' (image/audio) is not supported in v1 — see the roadmap. "
            f"v1 supports {SUPPORTED_SUMMARY}."
        )
    if ext in ARCHIVE_EXTENSIONS:
        raise UnsupportedModalityError(
            f"'{ext}' archives are not supported. v1 supports {SUPPORTED_SUMMARY}."
        )
    raise UnsupportedModalityError(
        f"Unknown file extension '{ext}'. v1 supports {SUPPORTED_SUMMARY}."
    )


def _sniff_magic(path: Path) -> str:
    """Return 'pdf', 'zip', or 'other' from the file's leading bytes."""
    with open(path, "rb") as handle:
        head = handle.read(4)
    if head.startswith(_PDF_MAGIC):
        return "pdf"
    if head.startswith(_ZIP_MAGIC):
        return "zip"
    return "other"


def _check_content_matches_extension(path: Path, file_type: str) -> None:
    """Catch wrong-extension artifacts via content signatures.

    A file whose content carries a *different* known format's signature than
    its extension implies raises :class:`ModalityMismatchError`. Content that
    merely fails to parse under the right signature is an
    :class:`InvalidArtifactError` (raised later by the format reader).
    """
    magic = _sniff_magic(path)
    if file_type == "pdf":
        if magic == "zip":
            raise ModalityMismatchError(
                f"'{path.name}' has a .pdf extension but its content looks like "
                "an Office/zip container."
            )
    elif file_type == "xlsx":
        if magic == "pdf":
            raise ModalityMismatchError(
                f"'{path.name}' has a .xlsx extension but its content looks like a PDF."
            )
    else:  # text-based: csv/txt/md
        if magic in ("pdf", "zip"):
            raise ModalityMismatchError(
                f"'{path.name}' has a .{file_type} extension but its content looks "
                f"like a binary {'PDF' if magic == 'pdf' else 'Office/zip'} file."
            )


@dataclass
class LoadedArtifact:
    """A successfully loaded artifact."""

    content: str
    file_type: str  # csv / xlsx / txt / md / pdf
    metadata: dict[str, Any] = field(default_factory=dict)


def read_file(path: str | Path, max_chars: int = 50000) -> LoadedArtifact:
    """Read a v1-supported file and return its textual content.

    Raises:
        InvalidArtifactError: missing, empty, or unparseable file.
        UnsupportedModalityError: extension outside the v1 set.
        ModalityMismatchError: content signature contradicts the extension.
        OptionalDependencyError: PDF/XLSX reader dependency missing.
    """
    p = Path(path)
    if not p.exists():
        raise InvalidArtifactError(f"File not found: {p}")

    file_type = file_type_for(p)

    if p.stat().st_size == 0:
        raise InvalidArtifactError(f"File is empty: {p.name}")

    _check_content_matches_extension(p, file_type)

    if file_type == "pdf":
        return _read_pdf(p, max_chars)
    if file_type == "xlsx":
        return _read_excel(p, max_chars)
    return _read_text(p, file_type, max_chars)


def _read_text(path: Path, file_type: str, max_chars: int) -> LoadedArtifact:
    """Read a text-based file (csv/txt/md)."""
    with open(path, encoding="utf-8", errors="replace") as handle:
        content = handle.read(max_chars)
    return LoadedArtifact(
        content=content,
        file_type=file_type,
        metadata={"truncated": len(content) == max_chars, "file_name": path.name},
    )


def _read_pdf(path: Path, max_chars: int) -> LoadedArtifact:
    """Read a PDF via PyMuPDF, falling back to pdfplumber."""
    if HAS_PYMUPDF:
        return _read_pdf_pymupdf(path, max_chars)
    if HAS_PDFPLUMBER:
        return _read_pdf_pdfplumber(path, max_chars)
    raise OptionalDependencyError(
        'Reading PDFs requires PyMuPDF or pdfplumber. Install with: pip install "traxr[document]"'
    )


def _read_pdf_pymupdf(path: Path, max_chars: int) -> LoadedArtifact:
    try:
        doc = fitz.open(path)
    except Exception as exc:
        raise InvalidArtifactError(f"Cannot open PDF '{path.name}': {exc}") from exc

    try:
        text_parts = []
        total_chars = 0
        num_pages = doc.page_count
        for page_no, page in enumerate(doc):
            page_text = page.get_text()
            if total_chars + len(page_text) > max_chars:
                remaining = max_chars - total_chars
                text_parts.append(f"--- Page {page_no + 1} ---\n{page_text[:remaining]}")
                total_chars = max_chars
                break
            text_parts.append(f"--- Page {page_no + 1} ---\n{page_text}")
            total_chars += len(page_text)
    except Exception as exc:
        raise InvalidArtifactError(f"Cannot read PDF '{path.name}': {exc}") from exc
    finally:
        doc.close()

    return LoadedArtifact(
        content="\n\n".join(text_parts),
        file_type="pdf",
        metadata={
            "num_pages": num_pages,
            "truncated": total_chars >= max_chars,
            "file_name": path.name,
        },
    )


def _read_pdf_pdfplumber(path: Path, max_chars: int) -> LoadedArtifact:
    text_parts = []
    total_chars = 0
    try:
        with pdfplumber.open(path) as pdf:
            num_pages = len(pdf.pages)
            for page_no, page in enumerate(pdf.pages):
                page_text = page.extract_text() or ""
                if total_chars + len(page_text) > max_chars:
                    remaining = max_chars - total_chars
                    text_parts.append(f"--- Page {page_no + 1} ---\n{page_text[:remaining]}")
                    total_chars = max_chars
                    break
                text_parts.append(f"--- Page {page_no + 1} ---\n{page_text}")
                total_chars += len(page_text)
    except Exception as exc:
        raise InvalidArtifactError(f"Cannot read PDF '{path.name}': {exc}") from exc

    return LoadedArtifact(
        content="\n\n".join(text_parts),
        file_type="pdf",
        metadata={
            "num_pages": num_pages,
            "truncated": total_chars >= max_chars,
            "file_name": path.name,
        },
    )


def _read_excel(path: Path, max_chars: int) -> LoadedArtifact:
    """Read an XLSX workbook as tab-separated text (one section per sheet)."""
    if not HAS_OPENPYXL:
        raise OptionalDependencyError(
            'Reading XLSX files requires openpyxl. Install with: pip install "traxr[document]"'
        )

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:
        raise InvalidArtifactError(f"Cannot open XLSX '{path.name}': {exc}") from exc

    try:
        lines = []
        sheet_names = list(wb.sheetnames)  # read before close() in the finally
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            for row in sheet.iter_rows():
                row_parts = []
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    row_parts.append(val)
                lines.append("\t".join(row_parts))
    finally:
        wb.close()

    content = "\n".join(lines)
    return LoadedArtifact(
        content=content[:max_chars],
        file_type="xlsx",
        metadata={
            "sheet_names": sheet_names,
            "truncated": len(content) > max_chars,
            "file_name": path.name,
        },
    )


# =========================================================================
# File inspection (pre-planning schema discovery)
# =========================================================================


@dataclass
class FileInspection:
    """Structural metadata of an artifact (v1 subset of the source dataclass)."""

    file_name: str
    file_type: str  # csv / xlsx / txt / md / pdf
    file_size: int
    mime_type: str = ""

    # Tabular structure
    columns: list[str] = field(default_factory=list)
    row_count: int = 0
    sample_rows: list[list[Any]] = field(default_factory=list)
    sheet_names: list[str] = field(default_factory=list)

    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        """Convert inspection results to a JSON-serializable dictionary."""
        result: dict[str, Any] = {
            "file_name": self.file_name,
            "file_type": self.file_type,
            "file_size": self.file_size,
            "mime_type": self.mime_type,
        }
        if self.columns:
            result["columns"] = self.columns
            result["row_count"] = self.row_count
            result["sample_rows"] = self.sample_rows
        if self.sheet_names:
            result["sheet_names"] = self.sheet_names
        if self.errors:
            result["errors"] = self.errors
        return result


def inspect_file(path: str | Path) -> FileInspection:
    """Inspect a v1-supported file's structure (columns, rows, sheets).

    Inspection is best-effort: structural probing failures are recorded in
    ``FileInspection.errors`` rather than raised. Unsupported extensions still
    raise :class:`UnsupportedModalityError` (fail-fast, like ``read_file``).
    """
    p = Path(path)
    file_type = file_type_for(p)
    if not p.exists():
        raise InvalidArtifactError(f"File not found: {p}")

    mime_type, _ = mimetypes.guess_type(str(p))
    inspection = FileInspection(
        file_name=p.name,
        file_type=file_type,
        file_size=p.stat().st_size,
        mime_type=mime_type or "",
    )

    try:
        if file_type == "csv":
            _inspect_csv(p, inspection)
        elif file_type == "xlsx":
            _inspect_excel(p, inspection)
    except Exception as exc:  # best-effort, mirrors source behavior
        inspection.errors.append(f"Inspection error: {exc}")

    return inspection


def _inspect_csv(path: Path, inspection: FileInspection) -> None:
    """Inspect CSV file structure (first 20 lines)."""
    with open(path, encoding="utf-8", errors="replace") as handle:
        lines = []
        for i, line in enumerate(handle):
            if i >= 20:
                break
            lines.append(line)

    if not lines:
        inspection.errors.append("Empty CSV file")
        return

    rows = list(csv.reader(lines))
    if rows:
        inspection.columns = [str(c).strip() for c in rows[0]]
        inspection.sample_rows = [list(r) for r in rows[1:4]] if len(rows) > 1 else []
        avg_line_len = sum(len(line) for line in lines) / len(lines)
        inspection.row_count = int(inspection.file_size / avg_line_len)


def _inspect_excel(path: Path, inspection: FileInspection) -> None:
    """Inspect XLSX file structure (first sheet, first 5 rows)."""
    if not HAS_OPENPYXL:
        inspection.errors.append("openpyxl not installed - cannot inspect Excel")
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        inspection.sheet_names = list(wb.sheetnames)
        ws = wb.active
        if ws is not None:
            inspection.row_count = ws.max_row or 0
            for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
                if i == 0:
                    inspection.columns = [str(c) if c is not None else "" for c in row]
                else:
                    inspection.sample_rows.append(list(row))
    finally:
        wb.close()
