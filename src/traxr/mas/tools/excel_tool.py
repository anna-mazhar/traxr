"""Excel tool for structured spreadsheet operations using pandas.

Enhanced with advanced capabilities for handling complex, messy, and
color-coded Excel files commonly found in real-world benchmarks.
"""

import pandas as pd
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from .base import BaseTool, ToolResult

import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Color
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelTool(BaseTool):
    """Tool for Excel file operations with structured DataFrame access.

    Provides deterministic operations for loading, inspecting, and querying
    spreadsheet data. Designed to handle messy real-world Excel files including:
    - Multi-header tables with metadata rows above the data
    - Blank separator rows between sections
    - Merged cells
    - Hidden rows/columns
    - Color-coded cells carrying semantic meaning
    - Multiple sheets
    - Named ranges and comments
    """

    def __init__(self, file_path: Optional[str] = None, corruption_fn=None):
        """Initialize ExcelTool.

        Args:
            file_path: Path to Excel file (can be set later via load)
            corruption_fn: Optional corruption function applied to DataFrame
        """
        super().__init__(name="excel")
        self.file_path = file_path
        self.df: Optional[pd.DataFrame] = None
        self.df_colors: Optional[pd.DataFrame] = None
        self.sheet_names: List[str] = []
        self.corruption_fn = corruption_fn
        self._workbook_metadata: Dict[str, Any] = {}

    def execute(self, operation: str, **kwargs) -> ToolResult:
        """Execute Excel operation."""
        operations = {
            "load": self._load,
            "load_sheet": self._load_sheet,
            "get_schema": self._get_schema,
            "head": self._head,
            "tail": self._tail,
            "filter": self._filter,
            "sort": self._sort,
            "query": self._query,
            "get_value": self._get_value,
            "get_column": self._get_column,
            "get_unique": self._get_unique,
            "describe": self._describe,
            "get_cell_color": self._get_cell_color,
            "get_colors_grid": self._get_colors_grid,
            "get_sheet_names": self._get_sheet_names,
            "get_merged_cells": self._get_merged_cells,
            "get_hidden_rows_cols": self._get_hidden_rows_cols,
            "get_cell_metadata": self._get_cell_metadata,
            "get_structural_summary": self._get_structural_summary,
            "get_named_ranges": self._get_named_ranges,
            "read_range": self._read_range,
        }

        if operation not in operations:
            return ToolResult(
                success=False,
                output=None,
                error=f"Unknown operation '{operation}'. Available: {list(operations.keys())}"
            )

        try:
            return operations[operation](**kwargs)
        except Exception as e:
            return ToolResult(
                success=False,
                output=None,
                error=f"Error in {operation}: {str(e)}"
            )

    def get_available_operations(self) -> List[str]:
        """Get list of available operations."""
        return [
            "load", "load_sheet", "get_schema", "head", "tail", "filter",
            "sort", "query", "get_value", "get_column", "get_unique",
            "describe", "get_cell_color", "get_colors_grid",
            "get_sheet_names", "get_merged_cells", "get_hidden_rows_cols",
            "get_cell_metadata", "get_structural_summary", "get_named_ranges",
            "read_range",
        ]

    def get_schema(self) -> "ToolSchema":
        from .tool_schema import ToolSchema, OperationSchema, ToolParameterSchema
        return ToolSchema(
            name="excel",
            description="Excel/spreadsheet tool for loading, inspecting, filtering, and querying tabular data. Handles messy files with merged cells, colors, hidden rows, and multi-sheet workbooks.",
            operations={
                "load": OperationSchema(
                    name="load",
                    description="Load an Excel file into a DataFrame. Auto-detects headers.",
                    parameters=[
                        ToolParameterSchema(name="file_path", type="string", description="Path to Excel file", required=False),
                        ToolParameterSchema(name="sheet_name", type="string", description="Sheet name or index to load (default: first sheet)", required=False, default="0"),
                    ],
                ),
                "load_sheet": OperationSchema(
                    name="load_sheet",
                    description="Load a different sheet from the same workbook.",
                    parameters=[
                        ToolParameterSchema(name="sheet_name", type="string", description="Sheet name or index"),
                    ],
                ),
                "get_schema": OperationSchema(
                    name="get_schema",
                    description="Get DataFrame schema (column names, data types, shape).",
                    parameters=[],
                ),
                "head": OperationSchema(
                    name="head",
                    description="Get the first N rows of the DataFrame.",
                    parameters=[
                        ToolParameterSchema(name="n", type="integer", description="Number of rows", required=False, default=5),
                    ],
                ),
                "tail": OperationSchema(
                    name="tail",
                    description="Get the last N rows of the DataFrame.",
                    parameters=[
                        ToolParameterSchema(name="n", type="integer", description="Number of rows", required=False, default=5),
                    ],
                ),
                "filter": OperationSchema(
                    name="filter",
                    description="Filter rows using pandas query syntax (e.g. 'Age > 30 and City == \"NYC\"').",
                    parameters=[
                        ToolParameterSchema(name="condition", type="string", description="Pandas query condition string"),
                    ],
                ),
                "sort": OperationSchema(
                    name="sort",
                    description="Sort DataFrame by one or more columns.",
                    parameters=[
                        ToolParameterSchema(name="by", type="string", description="Column name or list of column names to sort by"),
                        ToolParameterSchema(name="ascending", type="boolean", description="Sort ascending (true) or descending (false)", required=False, default=True),
                    ],
                ),
                "query": OperationSchema(
                    name="query",
                    description="Execute a pandas query expression.",
                    parameters=[
                        ToolParameterSchema(name="query", type="string", description="Pandas query expression"),
                    ],
                ),
                "get_value": OperationSchema(
                    name="get_value",
                    description="Get the value of a specific cell by row index and column name.",
                    parameters=[
                        ToolParameterSchema(name="row", type="integer", description="Row index (0-based)"),
                        ToolParameterSchema(name="column", type="string", description="Column name"),
                    ],
                ),
                "get_column": OperationSchema(
                    name="get_column",
                    description="Get all values from a column.",
                    parameters=[
                        ToolParameterSchema(name="column", type="string", description="Column name"),
                    ],
                ),
                "get_unique": OperationSchema(
                    name="get_unique",
                    description="Get unique values from a column.",
                    parameters=[
                        ToolParameterSchema(name="column", type="string", description="Column name"),
                    ],
                ),
                "describe": OperationSchema(
                    name="describe",
                    description="Get statistical summary of the DataFrame (count, mean, std, min, max, etc.).",
                    parameters=[],
                ),
                "get_cell_color": OperationSchema(
                    name="get_cell_color",
                    description="Get the background color of a specific cell as hex code.",
                    parameters=[
                        ToolParameterSchema(name="row", type="integer", description="Row index (0-based)"),
                        ToolParameterSchema(name="col", type="integer", description="Column index (0-based)"),
                    ],
                ),
                "get_colors_grid": OperationSchema(
                    name="get_colors_grid",
                    description="Load all cell background colors as a grid.",
                    parameters=[
                        ToolParameterSchema(name="max_rows", type="integer", description="Maximum rows to read", required=False),
                        ToolParameterSchema(name="max_cols", type="integer", description="Maximum columns to read", required=False),
                    ],
                ),
                "get_sheet_names": OperationSchema(
                    name="get_sheet_names",
                    description="Get list of all sheet names with their dimensions.",
                    parameters=[],
                ),
                "get_merged_cells": OperationSchema(
                    name="get_merged_cells",
                    description="Get all merged cell ranges in the active sheet.",
                    parameters=[],
                ),
                "get_hidden_rows_cols": OperationSchema(
                    name="get_hidden_rows_cols",
                    description="Get hidden rows and columns in the active sheet.",
                    parameters=[],
                ),
                "get_cell_metadata": OperationSchema(
                    name="get_cell_metadata",
                    description="Get rich metadata for a cell (formula, comment, formatting, color).",
                    parameters=[
                        ToolParameterSchema(name="row", type="integer", description="Row index (0-based)"),
                        ToolParameterSchema(name="col", type="integer", description="Column index (0-based)"),
                    ],
                ),
                "get_structural_summary": OperationSchema(
                    name="get_structural_summary",
                    description="Get comprehensive structural summary: blank rows, header detection, section markers, color patterns, merged cells.",
                    parameters=[],
                ),
                "get_named_ranges": OperationSchema(
                    name="get_named_ranges",
                    description="Get named ranges defined in the workbook.",
                    parameters=[],
                ),
                "read_range": OperationSchema(
                    name="read_range",
                    description="Read a specific rectangular range from the DataFrame.",
                    parameters=[
                        ToolParameterSchema(name="start_row", type="integer", description="Starting row index (0-based, inclusive)"),
                        ToolParameterSchema(name="end_row", type="integer", description="Ending row index (0-based, exclusive)"),
                        ToolParameterSchema(name="start_col", type="integer", description="Starting column index (0-based, inclusive)", required=False, default=0),
                        ToolParameterSchema(name="end_col", type="integer", description="Ending column index (0-based, exclusive)", required=False),
                    ],
                ),
            },
        )

    def _load(self, file_path: Optional[str] = None, sheet_name: Union[str, int] = 0) -> ToolResult:
        """Load Excel file into DataFrame.

        Args:
            file_path: Path to Excel file (optional if set in __init__)
            sheet_name: Sheet name or index to load (default: first sheet)

        Returns:
            ToolResult with DataFrame shape info
        """
        if file_path:
            self.file_path = file_path

        if not self.file_path:
            return ToolResult(
                success=False,
                output=None,
                error="No file_path provided"
            )

        # Determine engine based on file extension
        ext = Path(self.file_path).suffix.lower()
        engine = None
        if ext == '.xls':
            engine = 'xlrd'  # Old .xls format requires xlrd
        elif ext == '.xlsx':
            engine = 'openpyxl'  # Modern .xlsx format uses openpyxl

        # Try to load with appropriate engine, with fallback handling
        try:
            excel_file = pd.ExcelFile(self.file_path, engine=engine)
            self.sheet_names = excel_file.sheet_names
        except ImportError as ie:
            # Engine not installed - provide helpful error
            if 'xlrd' in str(ie):
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"xlrd not installed for .xls files. Install with: pip install xlrd"
                )
            elif 'openpyxl' in str(ie):
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"openpyxl not installed for .xlsx files. Install with: pip install openpyxl"
                )
            raise
        except Exception as e:
            # For .xls files, try without engine as fallback
            if ext == '.xls':
                try:
                    excel_file = pd.ExcelFile(self.file_path)
                    self.sheet_names = excel_file.sheet_names
                except Exception:
                    return ToolResult(
                        success=False,
                        output=None,
                        error=f"Failed to load .xls file: {str(e)}. Install xlrd with: pip install xlrd"
                    )
            else:
                raise

        # First load without headers to inspect row 0
        try:
            df_raw = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None, engine=engine)
        except ImportError as ie:
            if 'xlrd' in str(ie):
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"xlrd not installed for .xls files. Install with: pip install xlrd"
                )
            raise
        except Exception as e:
            # Fallback: try without engine specification
            try:
                df_raw = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
            except Exception:
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"Failed to read Excel file: {str(e)}"
                )

        # Auto-detect if row 0 looks like headers
        # Headers typically: mostly strings, unique values, not numeric data
        use_header = self._detect_header_row(df_raw)

        if use_header:
            # Reload with row 0 as headers
            try:
                self.df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0, engine=engine)
            except Exception:
                # Fallback without engine
                self.df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0)
            logger.debug(f"[ExcelTool] Auto-detected headers in row 0: {list(self.df.columns)}")
        else:
            self.df = df_raw
            logger.debug(f"[ExcelTool] No headers detected, using integer columns: {list(self.df.columns)}")

        # Apply corruption if configured
        if self.corruption_fn:
            self.df = self.corruption_fn(self.df)

        # Extract workbook-level metadata
        self._extract_workbook_metadata()

        # Also load cell background colors if openpyxl is available
        colors_loaded = False
        if OPENPYXL_AVAILABLE:
            try:
                color_result = self._get_colors_grid()
                if color_result.success:
                    colors_loaded = True
                    logger.debug(f"[ExcelTool] Loaded colors grid: {self.df_colors.shape}")
                else:
                    logger.debug(f"[ExcelTool] Failed to load colors: {color_result.error}")
            except Exception as e:
                logger.debug(f"[ExcelTool] Exception loading colors: {e}")

        return ToolResult(
            success=True,
            output=f"Loaded sheet '{sheet_name}' with shape {self.df.shape}",
            metadata={
                "shape": self.df.shape,
                "columns": list(self.df.columns),
                "sheet_names": self.sheet_names,
                "dtypes": {str(col): str(dtype) for col, dtype in self.df.dtypes.items()},
                "colors_loaded": colors_loaded,
                "num_sheets": len(self.sheet_names),
            }
        )

    def _load_sheet(self, sheet_name: Union[str, int] = 0) -> ToolResult:
        """Load a different sheet from the same workbook.

        Args:
            sheet_name: Sheet name or index

        Returns:
            ToolResult with new sheet info
        """
        if not self.file_path:
            return ToolResult(success=False, output=None, error="No file loaded")

        return self._load(sheet_name=sheet_name)

    def _extract_workbook_metadata(self):
        """Extract structural metadata from the workbook using openpyxl."""
        if not OPENPYXL_AVAILABLE or not self.file_path:
            return

        try:
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active

            metadata = {
                "merged_cells": [],
                "hidden_rows": [],
                "hidden_cols": [],
                "has_comments": False,
                "named_ranges": [],
                "max_row_actual": ws.max_row,
                "max_col_actual": ws.max_column,
            }

            # Merged cells
            for merged_range in ws.merged_cells.ranges:
                metadata["merged_cells"].append({
                    "range": str(merged_range),
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col,
                    "value": ws.cell(merged_range.min_row, merged_range.min_col).value,
                })

            # Hidden rows
            for row_idx, dim in ws.row_dimensions.items():
                if dim.hidden:
                    metadata["hidden_rows"].append(row_idx)

            # Hidden columns
            for col_letter, dim in ws.column_dimensions.items():
                if dim.hidden:
                    metadata["hidden_cols"].append(col_letter)

            # Check for comments
            for row in ws.iter_rows(max_row=min(ws.max_row or 1, 100), max_col=min(ws.max_column or 1, 50)):
                for cell in row:
                    if cell.comment:
                        metadata["has_comments"] = True
                        break
                if metadata["has_comments"]:
                    break

            # Named ranges (handle different openpyxl versions)
            try:
                defined = getattr(wb.defined_names, 'definedName', None) or wb.defined_names.values()
                for name in defined:
                    metadata["named_ranges"].append({
                        "name": name.name,
                        "value": name.attr_text,
                    })
            except (AttributeError, TypeError):
                pass

            wb.close()
            self._workbook_metadata = metadata

        except Exception as e:
            logger.debug(f"[ExcelTool] Metadata extraction warning: {e}")
            self._workbook_metadata = {}

    def _get_schema(self) -> ToolResult:
        """Get DataFrame schema (columns and types)."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        schema = {
            "columns": list(self.df.columns),
            "dtypes": {str(col): str(dtype) for col, dtype in self.df.dtypes.items()},
            "shape": self.df.shape,
            "sheet_names": self.sheet_names,
        }

        return ToolResult(
            success=True,
            output=schema,
            metadata=schema
        )

    def _head(self, n: int = 5) -> ToolResult:
        """Get first N rows."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        head_df = self.df.head(n)
        return ToolResult(
            success=True,
            output=head_df.to_dict('records'),
            metadata={"rows_returned": len(head_df)}
        )

    def _tail(self, n: int = 5) -> ToolResult:
        """Get last N rows."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        tail_df = self.df.tail(n)
        return ToolResult(
            success=True,
            output=tail_df.to_dict('records'),
            metadata={"rows_returned": len(tail_df)}
        )

    def _filter(self, condition: str) -> ToolResult:
        """Filter rows using pandas query syntax."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        filtered_df = self.df.query(condition)
        return ToolResult(
            success=True,
            output=filtered_df.to_dict('records'),
            metadata={
                "rows_returned": len(filtered_df),
                "total_rows": len(self.df),
                "condition": condition
            }
        )

    def _sort(self, by: Union[str, List[str]], ascending: bool = True) -> ToolResult:
        """Sort DataFrame by column(s)."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        sorted_df = self.df.sort_values(by=by, ascending=ascending)
        return ToolResult(
            success=True,
            output=sorted_df.to_dict('records'),
            metadata={
                "rows_returned": len(sorted_df),
                "sorted_by": by,
                "ascending": ascending
            }
        )

    def _query(self, query: str) -> ToolResult:
        """Execute pandas query."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        result_df = self.df.query(query)
        return ToolResult(
            success=True,
            output=result_df.to_dict('records'),
            metadata={
                "rows_returned": len(result_df),
                "query": query
            }
        )

    def _get_value(self, row: int, column: str) -> ToolResult:
        """Get specific cell value."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        if column not in self.df.columns:
            return ToolResult(
                success=False,
                output=None,
                error=f"Column '{column}' not found. Available: {list(self.df.columns)}"
            )

        if row >= len(self.df):
            return ToolResult(
                success=False,
                output=None,
                error=f"Row {row} out of bounds (total rows: {len(self.df)})"
            )

        value = self.df.iloc[row][column]
        return ToolResult(
            success=True,
            output=value,
            metadata={"row": row, "column": column, "value_type": str(type(value).__name__)}
        )

    def _get_column(self, column: str) -> ToolResult:
        """Get all values from a column."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        if column not in self.df.columns:
            return ToolResult(
                success=False,
                output=None,
                error=f"Column '{column}' not found. Available: {list(self.df.columns)}"
            )

        values = self.df[column].tolist()
        return ToolResult(
            success=True,
            output=values,
            metadata={"column": column, "count": len(values)}
        )

    def _get_unique(self, column: str) -> ToolResult:
        """Get unique values from a column."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        if column not in self.df.columns:
            return ToolResult(
                success=False,
                output=None,
                error=f"Column '{column}' not found. Available: {list(self.df.columns)}"
            )

        unique_values = self.df[column].unique().tolist()
        return ToolResult(
            success=True,
            output=unique_values,
            metadata={"column": column, "unique_count": len(unique_values)}
        )

    def _describe(self) -> ToolResult:
        """Get statistical summary of DataFrame."""
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        description = self.df.describe(include='all').to_dict()
        return ToolResult(
            success=True,
            output=description,
            metadata={"shape": self.df.shape}
        )

    def _get_cell_color(self, row: int, col: int) -> ToolResult:
        """Get the background color of a specific cell as 6-digit hex code."""
        if not OPENPYXL_AVAILABLE:
            return ToolResult(
                success=False, output=None,
                error="openpyxl not available - install with: pip install openpyxl"
            )

        if not self.file_path:
            return ToolResult(success=False, output=None, error="No file loaded")

        try:
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active
            cell = ws.cell(row=row + 1, column=col + 1)
            color_hex = self._extract_cell_color(cell)
            wb.close()

            return ToolResult(success=True, output=color_hex, metadata={"row": row, "col": col})

        except Exception as e:
            return ToolResult(success=False, output=None, error=f"Error reading cell color: {str(e)}")

    def _extract_cell_color(self, cell) -> Optional[str]:
        """Extract background color from an openpyxl cell, handling ARGB, theme, and indexed colors."""
        try:
            if not cell.fill or not cell.fill.start_color:
                return None

            color = cell.fill.start_color

            # Try RGB first (most reliable)
            if hasattr(color, 'rgb') and color.rgb is not None:
                try:
                    rgb = str(color.rgb)
                except (TypeError, ValueError):
                    rgb = None

                if rgb and rgb not in ('00000000', '0', 'None'):
                    # Filter out openpyxl validation error strings
                    if 'Values must be' in rgb or len(rgb) > 8:
                        pass
                    else:
                        if len(rgb) == 8:
                            rgb = rgb[2:]  # Strip alpha from ARGB
                        if len(rgb) == 6 and rgb != '000000':
                            return rgb

            # Try indexed color
            if hasattr(color, 'indexed') and color.indexed is not None:
                try:
                    from openpyxl.styles.colors import COLOR_INDEX
                    idx = int(color.indexed)
                    if idx < len(COLOR_INDEX):
                        idx_color = COLOR_INDEX[idx]
                        if idx_color and len(str(idx_color)) >= 6:
                            hex_val = str(idx_color)[-6:]
                            if hex_val != '000000':
                                return hex_val
                except (ImportError, IndexError, TypeError, ValueError):
                    pass

            # Theme colors - return a description since exact resolution requires theme XML
            if hasattr(color, 'theme') and color.theme is not None:
                try:
                    tint = float(getattr(color, 'tint', 0) or 0)
                    theme = int(color.theme)
                    if theme != 0 or tint != 0:
                        return f"theme:{theme}:tint:{tint}"
                except (TypeError, ValueError):
                    pass

            return None
        except Exception:
            return None

    def _get_colors_grid(self, max_rows: int = None, max_cols: int = None) -> ToolResult:
        """Load all cell background colors as a DataFrame matching the data grid."""
        if not OPENPYXL_AVAILABLE:
            return ToolResult(
                success=False, output=None,
                error="openpyxl not available"
            )

        if not self.file_path:
            return ToolResult(success=False, output=None, error="No file loaded")

        try:
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active

            rows = int(max_rows) if max_rows is not None else int(ws.max_row) if ws.max_row is not None else 0
            cols = int(max_cols) if max_cols is not None else int(ws.max_column) if ws.max_column is not None else 0

            color_data = []
            for row_idx in range(rows):
                row_colors = []
                for col_idx in range(cols):
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                    color_hex = self._extract_cell_color(cell)
                    row_colors.append(color_hex)
                color_data.append(row_colors)

            wb.close()

            self.df_colors = pd.DataFrame(color_data)

            return ToolResult(
                success=True,
                output=self.df_colors,
                metadata={"shape": self.df_colors.shape}
            )

        except Exception as e:
            return ToolResult(success=False, output=None, error=f"Error reading colors: {str(e)}")

    def _get_sheet_names(self) -> ToolResult:
        """Get list of all sheet names with their dimensions."""
        if not self.file_path:
            return ToolResult(success=False, output=None, error="No file loaded")

        if not OPENPYXL_AVAILABLE:
            return ToolResult(
                success=True,
                output=self.sheet_names,
                metadata={"count": len(self.sheet_names)}
            )

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            sheets_info = []
            for name in wb.sheetnames:
                ws = wb[name]
                sheets_info.append({
                    "name": name,
                    "dimensions": ws.dimensions,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                })
            wb.close()

            return ToolResult(
                success=True,
                output=sheets_info,
                metadata={"count": len(sheets_info)}
            )
        except Exception as e:
            return ToolResult(
                success=True,
                output=[{"name": n} for n in self.sheet_names],
                metadata={"count": len(self.sheet_names), "warning": str(e)}
            )

    def _get_merged_cells(self) -> ToolResult:
        """Get all merged cell ranges in the active sheet."""
        if not self._workbook_metadata:
            return ToolResult(success=True, output=[], metadata={"count": 0})

        merged = self._workbook_metadata.get("merged_cells", [])
        return ToolResult(
            success=True,
            output=merged,
            metadata={"count": len(merged)}
        )

    def _get_hidden_rows_cols(self) -> ToolResult:
        """Get hidden rows and columns in the active sheet."""
        if not self._workbook_metadata:
            return ToolResult(
                success=True,
                output={"hidden_rows": [], "hidden_cols": []},
                metadata={"hidden_rows_count": 0, "hidden_cols_count": 0}
            )

        hidden_rows = self._workbook_metadata.get("hidden_rows", [])
        hidden_cols = self._workbook_metadata.get("hidden_cols", [])

        return ToolResult(
            success=True,
            output={"hidden_rows": hidden_rows, "hidden_cols": hidden_cols},
            metadata={
                "hidden_rows_count": len(hidden_rows),
                "hidden_cols_count": len(hidden_cols),
            }
        )

    def _get_cell_metadata(self, row: int, col: int) -> ToolResult:
        """Get rich metadata for a specific cell (formula, comment, formatting).

        Args:
            row: Row index (0-based)
            col: Column index (0-based)
        """
        if not OPENPYXL_AVAILABLE or not self.file_path:
            return ToolResult(success=False, output=None, error="openpyxl or file not available")

        try:
            # Load with formulas
            wb_formulas = load_workbook(self.file_path, data_only=False)
            ws_formulas = wb_formulas.active
            cell_formula = ws_formulas.cell(row=row + 1, column=col + 1)

            # Load with values
            wb_values = load_workbook(self.file_path, data_only=True)
            ws_values = wb_values.active
            cell_value = ws_values.cell(row=row + 1, column=col + 1)

            metadata = {
                "value": cell_value.value,
                "formula": str(cell_formula.value) if cell_formula.value and str(cell_formula.value).startswith('=') else None,
                "number_format": cell_value.number_format,
                "is_date": cell_value.is_date if hasattr(cell_value, 'is_date') else False,
                "font_bold": cell_formula.font.bold if cell_formula.font else False,
                "font_italic": cell_formula.font.italic if cell_formula.font else False,
                "font_strikethrough": cell_formula.font.strikethrough if cell_formula.font else False,
                "bg_color": self._extract_cell_color(cell_formula),
                "comment": cell_formula.comment.text if cell_formula.comment else None,
            }

            wb_formulas.close()
            wb_values.close()

            return ToolResult(success=True, output=metadata, metadata={"row": row, "col": col})

        except Exception as e:
            return ToolResult(success=False, output=None, error=f"Error: {str(e)}")

    def _get_structural_summary(self) -> ToolResult:
        """Get a comprehensive structural summary of the spreadsheet.

        This analyzes the file for common patterns that affect data extraction:
        blank rows, potential header rows, section markers, color patterns, etc.
        """
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        summary = {}

        # Detect blank rows
        blank_rows = []
        for i in range(len(self.df)):
            row = self.df.iloc[i]
            if row.isna().all() or (row.astype(str).str.strip() == '').all():
                blank_rows.append(i)
        summary["blank_rows"] = blank_rows[:20]  # Cap for readability
        summary["blank_row_count"] = len(blank_rows)

        # Detect potential header row (first row where most cells are non-null strings)
        potential_headers = []
        for i in range(min(20, len(self.df))):
            row = self.df.iloc[i]
            non_null = row.dropna()
            if len(non_null) >= len(self.df.columns) * 0.4:
                # Check if values look like headers (mostly strings, not numbers)
                str_count = sum(1 for v in non_null if isinstance(v, str) and not v.replace('.', '').replace('-', '').isdigit())
                if str_count >= len(non_null) * 0.5:
                    potential_headers.append({
                        "row": i,
                        "values": [str(v) for v in row.tolist()],
                        "non_null_count": len(non_null),
                    })
        summary["potential_header_rows"] = potential_headers[:5]

        # Detect section markers (rows where only 1-2 cells have values)
        section_markers = []
        for i in range(len(self.df)):
            row = self.df.iloc[i]
            non_null = row.dropna()
            non_empty = non_null[non_null.astype(str).str.strip() != '']
            if 1 <= len(non_empty) <= 2 and len(non_empty) < len(self.df.columns) * 0.3:
                vals = [str(v) for v in non_empty.tolist()]
                # Skip purely numeric single values (likely data, not markers)
                if not all(v.replace('.', '').replace('-', '').isdigit() for v in vals):
                    section_markers.append({"row": i, "values": vals})
        summary["section_markers"] = section_markers[:20]

        # Color patterns summary
        if self.df_colors is not None:
            unique_colors = set()
            color_row_counts = {}
            for i in range(len(self.df_colors)):
                row_colors = self.df_colors.iloc[i].dropna().unique().tolist()
                for c in row_colors:
                    if c and c != 'None':
                        unique_colors.add(c)
                        if c not in color_row_counts:
                            color_row_counts[c] = 0
                        color_row_counts[c] += 1

            summary["unique_colors"] = list(unique_colors)[:20]
            summary["color_distribution"] = {c: count for c, count in sorted(color_row_counts.items(), key=lambda x: -x[1])[:10]}
            summary["has_meaningful_colors"] = len(unique_colors) > 0
        else:
            summary["has_meaningful_colors"] = False

        # Merged cells from metadata
        summary["merged_cell_count"] = len(self._workbook_metadata.get("merged_cells", []))
        summary["has_hidden_rows"] = len(self._workbook_metadata.get("hidden_rows", [])) > 0
        summary["has_hidden_cols"] = len(self._workbook_metadata.get("hidden_cols", [])) > 0
        summary["has_comments"] = self._workbook_metadata.get("has_comments", False)
        summary["num_sheets"] = len(self.sheet_names)
        summary["sheet_names"] = self.sheet_names

        # Detect phantom rows (rows at end that are all NaN)
        phantom_count = 0
        for i in range(len(self.df) - 1, -1, -1):
            row = self.df.iloc[i]
            if row.isna().all() or (row.astype(str).str.strip() == '').all():
                phantom_count += 1
            else:
                break
        summary["trailing_empty_rows"] = phantom_count

        # Data type analysis per column
        col_analysis = {}
        for col in self.df.columns:
            values = self.df[col].dropna()
            if len(values) == 0:
                col_analysis[str(col)] = "all_null"
                continue
            numeric_count = sum(1 for v in values if self._is_numeric(v))
            str_count = len(values) - numeric_count
            col_analysis[str(col)] = {
                "non_null": len(values),
                "numeric": numeric_count,
                "string": str_count,
                "sample": [str(v) for v in values.head(3).tolist()],
            }
        summary["column_analysis"] = col_analysis

        return ToolResult(
            success=True,
            output=summary,
            metadata={"analyzed_rows": len(self.df), "analyzed_cols": len(self.df.columns)}
        )

    def _is_numeric(self, value) -> bool:
        """Check if a value is numeric (int, float, or numeric string)."""
        if isinstance(value, (int, float)):
            return True
        try:
            float(str(value).replace(',', ''))
            return True
        except (ValueError, TypeError):
            return False

    def _detect_header_row(self, df: pd.DataFrame) -> bool:
        """Detect if row 0 looks like a header row.

        Headers typically have:
        - Mostly string values (not numbers)
        - Mostly unique values
        - Non-empty cells

        Returns:
            True if row 0 looks like headers, False otherwise
        """
        if len(df) == 0 or len(df.columns) == 0:
            return False

        row0 = df.iloc[0]
        non_null = row0.dropna()

        if len(non_null) == 0:
            return False

        # Check what fraction are strings (not numeric)
        string_count = 0
        for val in non_null:
            if isinstance(val, str) and not self._is_numeric(val):
                string_count += 1

        string_ratio = string_count / len(non_null)

        # Check uniqueness
        unique_ratio = len(non_null.unique()) / len(non_null) if len(non_null) > 0 else 0

        # Check coverage (how many columns have values in row 0)
        coverage = len(non_null) / len(df.columns) if len(df.columns) > 0 else 0

        # Row 0 is likely a header if:
        # - Most values are strings (>50%)
        # - Values are mostly unique (>70%)
        # - Good coverage (>50% of columns have values)
        is_header = string_ratio >= 0.5 and unique_ratio >= 0.7 and coverage >= 0.5

        return is_header

    def _get_named_ranges(self) -> ToolResult:
        """Get named ranges defined in the workbook."""
        named = self._workbook_metadata.get("named_ranges", [])
        return ToolResult(
            success=True,
            output=named,
            metadata={"count": len(named)}
        )

    def _read_range(self, start_row: int, end_row: int, start_col: int = 0, end_col: int = None) -> ToolResult:
        """Read a specific rectangular range from the DataFrame.

        Args:
            start_row: Starting row index (0-based, inclusive)
            end_row: Ending row index (0-based, exclusive)
            start_col: Starting column index (0-based, inclusive)
            end_col: Ending column index (0-based, exclusive). None means all columns.
        """
        if self.df is None:
            return ToolResult(success=False, output=None, error="No DataFrame loaded")

        if end_col is None:
            end_col = len(self.df.columns)

        subset = self.df.iloc[start_row:end_row, start_col:end_col]
        return ToolResult(
            success=True,
            output=subset.to_dict('records'),
            metadata={
                "rows_returned": len(subset),
                "cols_returned": len(subset.columns),
                "range": f"[{start_row}:{end_row}, {start_col}:{end_col}]",
            }
        )
