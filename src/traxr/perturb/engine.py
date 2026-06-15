"""Perturbation engine for applying and logging perturbations.

A missing-``openpyxl`` ``ImportError`` is raised as
:class:`traxr.errors.OptionalDependencyError` naming the ``[document]`` extra.
"""

from pathlib import Path

from traxr.errors import OptionalDependencyError

from .audio import AudioPerturbator
from .image import ImagePerturbator
from .pdf import PDFPerturbator
from .tabular import TabularPerturbator, lossless_number
from .types import PerturbationResult, PerturbationType


class PerturbationEngine:
    """Central engine for applying perturbations to file content.

    Automatically selects the appropriate perturbator based on file type
    and logs all perturbations for reproducibility.

    Usage:
        engine = PerturbationEngine(seed=42)

        # Apply single perturbation
        result = engine.apply(
            content="col1,col2\\n1,2\\n3,4",
            file_type="csv",
            perturbation=PerturbationType.COLUMN_SWAP,
        )

        # Get supported perturbations for a file type
        supported = engine.get_supported_perturbations("xlsx")
    """

    def __init__(self, seed: int = 42) -> None:
        """Initialize engine.

        Args:
            seed: Random seed for reproducible perturbations
        """
        self.seed = seed

        # Register perturbators
        self._tabular = TabularPerturbator(seed=seed)
        self._pdf = PDFPerturbator(seed=seed)
        self._image = ImagePerturbator(seed=seed)
        self._audio = AudioPerturbator(seed=seed)

        # Map file types to perturbators (text-based)
        self._type_handlers: dict[str, TabularPerturbator | PDFPerturbator] = {
            # Tabular types
            "csv": self._tabular,
            "xlsx": self._tabular,
            "xls": self._tabular,
            "excel": self._tabular,
            "tsv": self._tabular,
            "json": self._tabular,
            # PDF/document types
            "pdf": self._pdf,
            "txt": self._pdf,
            "text": self._pdf,
            "md": self._pdf,
            "markdown": self._pdf,
            "doc": self._pdf,
            "docx": self._pdf,
            "pptx": self._pdf,
        }

        # Map image file types (binary)
        self._image_handlers = {
            "png": self._image,
            "jpg": self._image,
            "jpeg": self._image,
            "gif": self._image,
            "webp": self._image,
            "bmp": self._image,
            "tiff": self._image,
        }

        # Map audio file types (binary)
        self._audio_handlers = {
            "mp3": self._audio,
            "wav": self._audio,
            "ogg": self._audio,
            "flac": self._audio,
            "m4a": self._audio,
            "aac": self._audio,
            "wma": self._audio,
        }

        # Track all applied perturbations
        self._history: list[PerturbationResult] = []

    def apply(
        self,
        content: str,
        file_type: str,
        perturbation: PerturbationType,
        file_name: str = "",
    ) -> PerturbationResult:
        """Apply a perturbation to content.

        Args:
            content: Raw file content
            file_type: File type (csv, xlsx, pdf, etc.)
            perturbation: Which perturbation to apply
            file_name: Original file name (for logging)

        Returns:
            PerturbationResult with corrupted content and metadata
        """
        file_type = file_type.lower().lstrip(".")

        # Handle NULL case universally
        if perturbation == PerturbationType.NULL_CONTENT:
            result = PerturbationResult(
                original_content=content,
                corrupted_content="",
                perturbation_type=perturbation,
                description="Content replaced with empty",
                file_type=file_type,
                file_name=file_name,
            )
            self._history.append(result)
            return result

        # Find appropriate handler
        handler = self._type_handlers.get(file_type)

        if handler is None:
            # No handler for this file type
            result = PerturbationResult(
                original_content=content,
                corrupted_content=content,
                perturbation_type=perturbation,
                description="",
                applied=False,
                skip_reason=f"No perturbator for file type: {file_type}",
                file_type=file_type,
                file_name=file_name,
            )
            self._history.append(result)
            return result

        # Apply perturbation
        result = handler.apply(
            content=content,
            perturbation=perturbation,
            file_type=file_type,
            file_name=file_name,
        )

        self._history.append(result)
        return result

    def apply_image(
        self,
        content: bytes,
        file_type: str,
        perturbation: PerturbationType,
        file_name: str = "",
    ) -> PerturbationResult:
        """Apply a perturbation to image content.

        Args:
            content: Raw image bytes
            file_type: File type (png, jpg, etc.)
            perturbation: Which perturbation to apply
            file_name: Original file name (for logging)

        Returns:
            PerturbationResult with corrupted image bytes in corrupted_bytes field
        """
        file_type = file_type.lower().lstrip(".")

        # Handle NULL case
        if perturbation == PerturbationType.NULL_CONTENT:
            result = PerturbationResult(
                original_content="[binary image data]",
                corrupted_content="",
                perturbation_type=perturbation,
                description="Image replaced with empty content",
                file_type=file_type,
                file_name=file_name,
                corrupted_bytes=b"",
            )
            self._history.append(result)
            return result

        # Find appropriate handler
        image_handler = self._image_handlers.get(file_type)

        if image_handler is None:
            result = PerturbationResult(
                original_content="[binary image data]",
                corrupted_content="",
                perturbation_type=perturbation,
                description="",
                applied=False,
                skip_reason=f"No image perturbator for file type: {file_type}",
                file_type=file_type,
                file_name=file_name,
            )
            self._history.append(result)
            return result

        # Apply perturbation
        result = image_handler.apply(
            content=content,
            perturbation=perturbation,
            file_type=file_type,
            file_name=file_name,
        )

        self._history.append(result)
        return result

    def is_image_type(self, file_type: str) -> bool:
        """Check if file type is an image type."""
        return file_type.lower().lstrip(".") in self._image_handlers

    def apply_audio(
        self,
        content: bytes,
        file_type: str,
        perturbation: PerturbationType,
        file_name: str = "",
    ) -> PerturbationResult:
        """Apply a perturbation to audio content.

        Args:
            content: Raw audio bytes
            file_type: File type (mp3, wav, etc.)
            perturbation: Which perturbation to apply
            file_name: Original file name (for logging)

        Returns:
            PerturbationResult with corrupted audio bytes in corrupted_bytes field
        """
        file_type = file_type.lower().lstrip(".")

        # Handle NULL case
        if perturbation == PerturbationType.NULL_CONTENT:
            result = PerturbationResult(
                original_content="[binary audio data]",
                corrupted_content="",
                perturbation_type=perturbation,
                description="Audio replaced with empty content",
                file_type=file_type,
                file_name=file_name,
                corrupted_bytes=b"",
            )
            self._history.append(result)
            return result

        # Find appropriate handler
        audio_handler = self._audio_handlers.get(file_type)

        if audio_handler is None:
            result = PerturbationResult(
                original_content="[binary audio data]",
                corrupted_content="",
                perturbation_type=perturbation,
                description="",
                applied=False,
                skip_reason=f"No audio perturbator for file type: {file_type}",
                file_type=file_type,
                file_name=file_name,
            )
            self._history.append(result)
            return result

        # Apply perturbation
        result = audio_handler.apply(
            content=content,
            perturbation=perturbation,
            file_type=file_type,
            file_name=file_name,
        )

        self._history.append(result)
        return result

    def is_audio_type(self, file_type: str) -> bool:
        """Check if file type is an audio type."""
        return file_type.lower().lstrip(".") in self._audio_handlers

    def is_zip_type(self, file_type: str) -> bool:
        """Check if file type is a zip archive."""
        return file_type.lower().lstrip(".") in {"zip", "tar", "gz", "7z", "rar"}

    def apply_from_file(
        self,
        file_path: str,
        perturbation: PerturbationType,
    ) -> PerturbationResult:
        """Apply perturbation to a file.

        Args:
            file_path: Path to file
            perturbation: Which perturbation to apply

        Returns:
            PerturbationResult with corrupted content
        """
        path = Path(file_path)
        file_type = path.suffix.lower().lstrip(".")
        file_name = path.name

        # Read content
        content = self._read_file(file_path, file_type)

        return self.apply(
            content=content,
            file_type=file_type,
            perturbation=perturbation,
            file_name=file_name,
        )

    def get_supported_perturbations(self, file_type: str) -> list[PerturbationType]:
        """Get perturbations supported for a file type.

        Args:
            file_type: File type to check

        Returns:
            List of supported PerturbationType values
        """
        file_type = file_type.lower().lstrip(".")

        # NULL is always supported
        supported = [PerturbationType.NULL_CONTENT]

        # Tabular file types
        tabular_types = {"csv", "xlsx", "xls", "excel", "tsv", "json"}
        # PDF/document file types
        pdf_types = {"pdf", "txt", "text", "md", "markdown", "doc", "docx", "pptx"}
        # Image file types
        image_types = {"png", "jpg", "jpeg", "gif", "webp", "bmp", "tiff"}
        # Audio file types
        audio_types = {"mp3", "wav", "ogg", "flac", "m4a", "aac", "wma"}
        # Archive file types (only NULL_CONTENT supported)
        archive_types = {"zip", "tar", "gz", "7z", "rar"}

        # Archive files only support NULL_CONTENT (already in supported)
        if file_type in archive_types:
            return supported

        # Add type-specific perturbations
        if file_type in tabular_types:
            supported.extend(
                [
                    PerturbationType.COLUMN_SWAP,
                    PerturbationType.LABEL_CORRUPT,
                    PerturbationType.DATA_TYPE_CORRUPT,
                    PerturbationType.ROW_DUPLICATE,
                    PerturbationType.IRRELEVANT_COLUMNS,
                    PerturbationType.UNIT_CHANGE,
                ]
            )
        elif file_type in pdf_types:
            supported.extend(
                [
                    PerturbationType.OCR_NOISE,
                    PerturbationType.NUMBER_CORRUPTION,
                    PerturbationType.TEXT_REDACTION,
                    PerturbationType.PARAGRAPH_SHUFFLE,
                    PerturbationType.ENCODING_ERROR,
                    PerturbationType.SECTION_REMOVAL,
                ]
            )
        elif file_type in image_types:
            supported.extend(
                [
                    PerturbationType.BLUR,
                    PerturbationType.NOISE,
                    PerturbationType.LOW_RESOLUTION,
                    PerturbationType.PARTIAL_OCCLUSION,
                    PerturbationType.CONTRAST_REDUCTION,
                    PerturbationType.WATERMARK,
                ]
            )
        elif file_type in audio_types:
            supported.extend(
                [
                    PerturbationType.BACKGROUND_NOISE,
                    PerturbationType.SPEED_CHANGE,
                    PerturbationType.LOW_PASS_FILTER,
                ]
            )

        return supported

    def get_history(self) -> list[PerturbationResult]:
        """Get history of all applied perturbations."""
        return self._history.copy()

    def clear_history(self) -> None:
        """Clear perturbation history."""
        self._history.clear()

    def _read_file(self, file_path: str, file_type: str) -> str:
        """Read file content.

        Args:
            file_path: Path to file
            file_type: File type hint

        Returns:
            File content as string
        """
        path = Path(file_path)

        if file_type in ("xlsx", "xls", "excel"):
            return self._read_excel(path)
        else:
            # Text-based files
            return path.read_text(encoding="utf-8", errors="replace")

    def _read_excel(self, path: Path) -> str:
        """Read Excel file as TSV text."""
        try:
            import openpyxl
        except ImportError as exc:
            raise OptionalDependencyError(
                'openpyxl is required for Excel files. Install with: pip install "traxr[document]"'
            ) from exc

        wb = openpyxl.load_workbook(path)
        lines = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")

            for row in sheet.iter_rows():
                row_parts = []
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    row_parts.append(val)
                lines.append("\t".join(row_parts))

        wb.close()
        return "\n".join(lines)

    def write_excel(self, content: str, output_path: str) -> None:
        """Write TSV/CSV content to an Excel file.

        Args:
            content: TSV or CSV content (as produced by perturbations)
            output_path: Path to write the Excel file
        """
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise OptionalDependencyError(
                'openpyxl is required for Excel files. Install with: pip install "traxr[document]"'
            ) from exc

        # Detect delimiter
        lines = content.split("\n")[:5]
        tab_count = sum(line.count("\t") for line in lines)
        comma_count = sum(line.count(",") for line in lines)
        delimiter = "\t" if tab_count > comma_count else ","

        # Parse the content
        import csv
        import io

        rows = []
        for line in content.split("\n"):
            # Skip sheet headers from Excel format
            if line.strip().startswith("===") and "===" in line:
                continue
            if line.strip():
                reader = csv.reader(io.StringIO(line), delimiter=delimiter)
                for row in reader:
                    rows.append(row)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        assert ws is not None  # a fresh Workbook always has an active sheet
        ws.title = "Sheet1"

        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                # Convert to a number only when the coercion is lossless, so the
                # perturbed file does not silently re-type cells (e.g. "007").
                ws.cell(row=row_idx, column=col_idx, value=lossless_number(value))

        wb.save(output_path)
        wb.close()


def get_all_perturbation_types() -> list[PerturbationType]:
    """Get all available perturbation types."""
    return list(PerturbationType)


def get_tabular_perturbation_types() -> list[PerturbationType]:
    """Get perturbation types for tabular data."""
    return [
        PerturbationType.COLUMN_SWAP,
        PerturbationType.LABEL_CORRUPT,
        PerturbationType.DATA_TYPE_CORRUPT,
        PerturbationType.ROW_DUPLICATE,
        PerturbationType.IRRELEVANT_COLUMNS,
        PerturbationType.UNIT_CHANGE,
        PerturbationType.NULL_CONTENT,
    ]


def get_pdf_perturbation_types() -> list[PerturbationType]:
    """Get perturbation types for PDF/document data."""
    return [
        PerturbationType.OCR_NOISE,
        PerturbationType.NUMBER_CORRUPTION,
        PerturbationType.TEXT_REDACTION,
        PerturbationType.PARAGRAPH_SHUFFLE,
        PerturbationType.ENCODING_ERROR,
        PerturbationType.SECTION_REMOVAL,
        PerturbationType.NULL_CONTENT,
    ]
