"""Category 1 tests: tabular perturbation operators (CSV/XLSX round-trip)."""

import csv
import io
from pathlib import Path

import pytest

from traxr.perturb import PerturbationEngine, PerturbationType

TABULAR_OPERATORS = [
    PerturbationType.COLUMN_SWAP,
    PerturbationType.LABEL_CORRUPT,
    PerturbationType.DATA_TYPE_CORRUPT,
    PerturbationType.ROW_DUPLICATE,
    PerturbationType.IRRELEVANT_COLUMNS,
    PerturbationType.UNIT_CHANGE,
]


@pytest.fixture()
def csv_content(fixtures_dir: Path) -> str:
    return (fixtures_dir / "sample.csv").read_text()


def parse_csv(content: str) -> list[list[str]]:
    return [row for row in csv.reader(io.StringIO(content)) if row]


@pytest.mark.parametrize("op", TABULAR_OPERATORS)
def test_determinism_same_seed_identical_hashes(op: PerturbationType, csv_content: str) -> None:
    results = [PerturbationEngine(seed=7).apply(csv_content, "csv", op) for _ in range(2)]
    assert results[0].corrupted_hash == results[1].corrupted_hash
    assert results[0].corrupted_content == results[1].corrupted_content
    # Repeated apply on the SAME engine instance is also deterministic
    # (RNG reset per apply).
    engine = PerturbationEngine(seed=7)
    again = [engine.apply(csv_content, "csv", op) for _ in range(2)]
    assert again[0].corrupted_hash == again[1].corrupted_hash


@pytest.mark.parametrize("op", TABULAR_OPERATORS)
def test_seed_variation_changes_output(op: PerturbationType, csv_content: str) -> None:
    outputs = {
        PerturbationEngine(seed=s).apply(csv_content, "csv", op).corrupted_hash for s in range(8)
    }
    assert len(outputs) > 1, f"{op.value} produced identical output across 8 seeds"


@pytest.mark.parametrize("op", TABULAR_OPERATORS)
def test_applied_with_changes_recorded(op: PerturbationType, csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(csv_content, "csv", op)
    assert result.applied
    assert result.skip_reason is None
    assert result.changes, f"{op.value} recorded no changes"
    assert result.content_changed
    assert result.description


@pytest.mark.parametrize("op", TABULAR_OPERATORS)
def test_round_trip_csv_stays_parseable(op: PerturbationType, csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(csv_content, "csv", op)
    rows = parse_csv(result.corrupted_content)
    assert len(rows) >= 2  # header + at least one data row survived
    header_len = len(rows[0])
    assert all(len(row) == header_len for row in rows)


@pytest.mark.parametrize("op", TABULAR_OPERATORS)
def test_skip_on_insufficient_data(op: PerturbationType) -> None:
    result = PerturbationEngine(seed=42).apply("name,amount", "csv", op)
    assert not result.applied
    assert result.skip_reason == "Insufficient data (need header + at least 1 row)"
    assert result.corrupted_content == "name,amount"


# ---------------------------------------------------------------------------
# Intensity / semantics sanity per operator
# ---------------------------------------------------------------------------


def test_column_swap_preserves_cell_multiset(csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(csv_content, "csv", PerturbationType.COLUMN_SWAP)
    original = parse_csv(csv_content)
    corrupted = parse_csv(result.corrupted_content)
    assert len(corrupted) == len(original)
    for orig_row, new_row in zip(original, corrupted, strict=True):
        assert sorted(orig_row) == sorted(new_row)
    (change,) = result.changes
    assert change["col1_idx"] != change["col2_idx"]


def test_label_corrupt_only_touches_header(csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(csv_content, "csv", PerturbationType.LABEL_CORRUPT)
    original = parse_csv(csv_content)
    corrupted = parse_csv(result.corrupted_content)
    assert corrupted[0] != original[0]
    assert corrupted[1:] == original[1:]
    for change in result.changes:
        assert corrupted[0][change["col_idx"]] == change["corrupted"]


def test_data_type_corrupt_decorates_numeric_cells(csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(
        csv_content, "csv", PerturbationType.DATA_TYPE_CORRUPT
    )
    corrupted = parse_csv(result.corrupted_content)
    decorated = [
        cell for row in corrupted[1:] for cell in row if cell.startswith("~") or cell.endswith("*")
    ]
    assert decorated  # every numeric cell is decorated; sample has many
    for change in result.changes:
        assert corrupted[change["row"]][change["col"]] == change["corrupted"]


def test_row_duplicate_adds_marked_rows(csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(csv_content, "csv", PerturbationType.ROW_DUPLICATE)
    original = parse_csv(csv_content)
    corrupted = parse_csv(result.corrupted_content)
    added = len(corrupted) - len(original)
    assert added == len(result.changes)
    assert 1 <= added <= len(original) - 1  # 30% chance per row; seeded


def test_irrelevant_columns_appends_noise_columns(csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(
        csv_content, "csv", PerturbationType.IRRELEVANT_COLUMNS
    )
    original = parse_csv(csv_content)
    corrupted = parse_csv(result.corrupted_content)
    added = len(corrupted[0]) - len(original[0])
    assert 1 <= added <= 2
    assert all(corrupted[0][len(original[0]) + i].startswith("_") for i in range(added))
    # Original columns untouched
    for orig_row, new_row in zip(original, corrupted, strict=True):
        assert new_row[: len(orig_row)] == orig_row


def test_unit_change_multiplies_one_column(csv_content: str) -> None:
    result = PerturbationEngine(seed=42).apply(csv_content, "csv", PerturbationType.UNIT_CHANGE)
    (change,) = result.changes
    original = parse_csv(csv_content)
    corrupted = parse_csv(result.corrupted_content)
    col = change["col_idx"]
    multiplier = change["multiplier"]
    for orig_row, new_row in zip(original[1:], corrupted[1:], strict=True):
        expected = float(orig_row[col]) * multiplier
        assert float(new_row[col]) == pytest.approx(expected, abs=0.01)


def test_unit_change_skips_without_numeric_columns() -> None:
    content = "name,city\nann,berlin\nbob,paris\n"
    result = PerturbationEngine(seed=42).apply(content, "csv", PerturbationType.UNIT_CHANGE)
    assert result.applied  # parseable, but nothing to change
    assert not result.content_changed
    assert result.changes == []


def test_xlsx_round_trip_through_engine(fixtures_dir: Path, tmp_path: Path) -> None:
    """XLSX is read as TSV text, perturbed, and written back to a real file."""
    engine = PerturbationEngine(seed=42)
    result = engine.apply_from_file(str(fixtures_dir / "sample.xlsx"), PerturbationType.COLUMN_SWAP)
    assert result.applied
    assert result.file_type == "xlsx"
    out = tmp_path / "sample.xlsx"
    engine.write_excel(result.corrupted_content, str(out))

    import openpyxl

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    wb.close()
    assert len(rows) >= 2
    (change,) = result.changes
    # The swapped headers landed in the file.
    assert rows[0][change["col1_idx"]] == change["col2_name"]
    assert rows[0][change["col2_idx"]] == change["col1_name"]


class TestLosslessNumericCoercion:
    """M4: output serialization must not silently re-type string cells."""

    def test_lossless_number_preserves_ambiguous_strings(self) -> None:
        from traxr.perturb.tabular import lossless_number

        for raw in ("007", "1_000", "+5", "1e3", "NaN", "inf", "-inf", "hello"):
            assert lossless_number(raw) == raw  # kept as text
        # Genuine canonical numbers still convert.
        assert lossless_number("42") == 42
        assert lossless_number("3.14") == 3.14
        assert lossless_number("-7") == -7

    def test_to_json_keeps_ambiguous_cells_as_strings_and_is_valid(self) -> None:
        import json

        from traxr.perturb.tabular import TabularPerturbator

        rows = [["id", "amt", "flag"], ["007", "3.14", "NaN"], ["42", "1_000", "x"]]
        out = TabularPerturbator(seed=1)._to_json(rows)
        parsed = json.loads(out)  # must be valid JSON (no bare NaN literal)
        assert parsed[0] == {"id": "007", "amt": 3.14, "flag": "NaN"}
        assert parsed[1] == {"id": 42, "amt": "1_000", "flag": "x"}

    def test_write_excel_keeps_leading_zero_string(self, tmp_path: Path) -> None:
        import openpyxl

        engine = PerturbationEngine(seed=1)
        out = tmp_path / "nums.xlsx"
        engine.write_excel("id\tamt\n007\t42\n", str(out))
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()
        assert rows[1][0] == "007"  # preserved as text, not 7
        assert rows[1][1] == 42  # genuine integer still numeric
